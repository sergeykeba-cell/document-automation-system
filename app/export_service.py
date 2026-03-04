"""
export_service.py — експорт реєстру в .xlsx
"""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.database import get_conn

SHEET_MAP = {
    "аналізи":        "Аналізи",
    "влк":            "ВЛК",
    "стаціонар":      "Стаціонар",
    "характеристика": "Характеристика",
    "рапорт":         "Рапорт",
}

HEADERS = [
    "№", "П.І.Б.", "Телефон", "Звання", "Дата народження",
    "Дата зарахування", "Розміщення о/с", "Діагноз",
    "Дата створення", "Шлях до файлу", "Створив",
]


def _style_header(ws):
    fill = PatternFill("solid", fgColor="1A1E29")
    font = Font(bold=True, color="94A3B8", size=10)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="left")
    ws.row_dimensions[1].height = 20


def export_registry(doc_type: str | None = None) -> bytes:
    """
    doc_type=None → всі аркуші в одному файлі
    doc_type='аналізи' → тільки цей аркуш
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # видалити порожній Sheet

    types = [doc_type] if doc_type else list(SHEET_MAP.keys())

    with get_conn() as conn:
        for t in types:
            ws = wb.create_sheet(SHEET_MAP[t])
            ws.append(HEADERS)
            _style_header(ws)

            rows = conn.execute("""
                SELECT
                    ROW_NUMBER() OVER (ORDER BY d.id) AS num,
                    p.pib, p.phone, p.rank, p.birth_date,
                    p.enroll_date, p.location,
                    d.diagnosis, d.created_at, d.file_path, d.created_by
                FROM documents d
                JOIN personnel p ON p.id = d.personnel_id
                WHERE d.doc_type = ?
                ORDER BY d.id DESC
            """, (t,)).fetchall()

            for row in rows:
                ws.append(list(row))

            # Ширина колонок
            col_widths = [5, 35, 18, 20, 15, 15, 20, 40, 12, 50, 15]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(i)
                ].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
